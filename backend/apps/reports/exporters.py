"""
Reports exporters — CSV and Excel generation for sales reports.
"""
import csv
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse

from apps.orders.models import Order, OrderItem


def export_orders_csv(orders_qs, filename: str = 'sales_report.csv') -> HttpResponse:
    """Export orders queryset as CSV with UTF-8 BOM for Excel compatibility."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM for Excel UTF-8

    writer = csv.writer(response)
    writer.writerow([
        'Дата', 'Номер', 'Устройство', 'Обект', 'Тип',
        'Подсума', 'Отстъпка', 'Общо', 'Статус',
    ])

    for order in orders_qs.select_related('device__location'):
        writer.writerow([
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.order_number,
            order.device.display_name if order.device else '',
            order.device.location.name if order.device and order.device.location else '',
            order.get_order_type_display(),
            order.subtotal,
            order.discount,
            order.total,
            order.get_status_display(),
        ])

    return response


def export_orders_excel(orders_qs, filename: str = 'sales_report.xlsx') -> HttpResponse:
    """Export orders queryset as Excel (.xlsx) using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Продажби'

    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_font = Font(name='Calibri', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Headers
    headers = [
        'Дата', 'Номер', 'Устройство', 'Обект', 'Тип',
        'Подсума', 'Отстъпка', 'Общо', 'Статус',
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row_idx = 2
    for order in orders_qs.select_related('device__location'):
        row_data = [
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.order_number,
            order.device.display_name if order.device else '',
            order.device.location.name if order.device and order.device.location else '',
            order.get_order_type_display(),
            float(order.subtotal),
            float(order.discount),
            float(order.total),
            order.get_status_display(),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx in (6, 7, 8):  # Numeric columns
                cell.number_format = '#,##0.00'
        row_idx += 1

    # Auto column width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, row_idx)
        ) if row_idx > 1 else len(headers[col_idx - 1])
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Freeze first row
    ws.freeze_panes = 'A2'

    # Write to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_order_items_excel(orders_qs, filename: str = 'items_report.xlsx') -> HttpResponse:
    """Export detailed order items as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Артикули'

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        'Дата', 'Поръчка', 'Артикул', 'Ед. цена', 'Кол-во',
        'Отстъпка %', 'ДДС група', 'Ред общо',
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    items = OrderItem.objects.filter(
        order__in=orders_qs,
    ).select_related('order')

    for item in items:
        row_data = [
            item.order.created_at.strftime('%Y-%m-%d %H:%M'),
            item.order.order_number,
            item.product_name,
            float(item.product_price),
            float(item.quantity),
            item.discount_pct,
            item.vat_group,
            float(item.line_total),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in (4, 5, 6, 8):
                cell.number_format = '#,##0.00'
        row_idx += 1

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_vat_breakdown_excel(vat_data: list, date_from: str, date_to: str,
                                filename: str = 'vat_report.xlsx') -> HttpResponse:
    """Export VAT breakdown as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ДДС разбивка'

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    # Title
    ws.cell(row=1, column=1, value=f'ДДС разбивка: {date_from} — {date_to}')
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    headers = ['ДДС група', 'Обща сума', 'Брой артикули']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(vat_data, 4):
        ws.cell(row=row_idx, column=1, value=item.get('vat_group', ''))
        ws.cell(row=row_idx, column=2, value=float(item.get('total_amount', 0)))
        ws.cell(row=row_idx, column=2).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=3, value=item.get('total_items', 0))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
